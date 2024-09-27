import pandas as pd
import datetime
from openpyxl import Workbook

def calculate_age(birthdate):
    today = datetime.date.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

try:
    # Читання CSV файлу
    try:
        df = pd.read_csv('people_data.csv', delimiter=';', encoding='utf-8')
    except FileNotFoundError:
        print("Помилка: файл CSV не знайдено.")
        exit()
    except pd.errors.ParserError:
        print("Помилка: проблема при відкритті CSV файлу.")
        exit()

    # Додавання колонки з віком
    df['Дата народження'] = pd.to_datetime(df['Дата народження'], errors='coerce')
    df['Вік'] = df['Дата народження'].apply(lambda x: calculate_age(x) if pd.notnull(x) else None)

    # Створення нового XLSX файлу
    try:
        wb = Workbook()

        # Аркуші
        ws_all = wb.active
        ws_all.title = "all"
        ws_younger_18 = wb.create_sheet(title="younger_18")
        ws_18_45 = wb.create_sheet(title="18-45")
        ws_45_70 = wb.create_sheet(title="45-70")
        ws_older_70 = wb.create_sheet(title="older_70")

        # Запис на аркуш "all"
        ws_all.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])
        for i, row in df.iterrows():
            ws_all.append([i+1, row['Прізвище'], row['Ім’я'], row['По батькові'], row['Дата народження'].date(), row['Вік']])

        # Запис даних на аркуші з відповідними віковими групами
        def write_age_group(sheet, min_age=None, max_age=None):
            sheet.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])
            for i, row in df.iterrows():
                if pd.isnull(row['Вік']):
                    continue
                if (min_age is None or row['Вік'] >= min_age) and (max_age is None or row['Вік'] < max_age):
                    sheet.append([i+1, row['Прізвище'], row['Ім’я'], row['По батькові'], row['Дата народження'].date(), row['Вік']])

        write_age_group(ws_younger_18, max_age=18)
        write_age_group(ws_18_45, min_age=18, max_age=45)
        write_age_group(ws_45_70, min_age=45, max_age=70)
        write_age_group(ws_older_70, min_age=70)

        # Збереження файлу
        wb.save('people_data.xlsx')
        print("Ok")
    except Exception as e:
        print(f"Помилка: не вдалося створити файл XLSX. {e}")

except Exception as e:
    print(f"Сталася помилка: {e}")